#!/usr/bin/env python3
"""
exportar_notas.py
Genera automáticamente la planilla de notas del corte desde Supabase.

Uso:
  python3 exportar_notas.py --corte 1
  python3 exportar_notas.py --corte 2 --peso-formativa 0.70
  python3 exportar_notas.py --corte 1 --parciales parciales_corte1.csv

El archivo de parciales (opcional) es un CSV simple con columnas: student_id,nota_parcial
Si no se pasa, la columna Nota Parcial queda en blanco (para llenar en Excel).
"""

import os, sys, json, argparse, urllib.request, urllib.error
from datetime import date

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Instala openpyxl: pip3 install openpyxl")
    sys.exit(1)

# ── Configuración ──────────────────────────────────────────────────────────────
CORTE_RANGES = {1: (1, 5), 2: (6, 10), 3: (11, 14)}

# Pesos dentro de la nota formativa
W_QUIZ = 0.50        # quiz promedio (0-10 → 0-5)
W_ACTIVIDADES = 0.30 # actividades completadas / esperadas
W_VISITAS = 0.20     # semanas visitadas / esperadas

NOTA_MIN_APROBACION = 3.0

# Colores IUB
IUB_NAVY   = "1E2843"
IUB_YELLOW = "FFDF2D"

def get_supabase_config():
    """Lee la config de supabase-config.js"""
    config_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), "js", "supabase-config.js")
    if not os.path.exists(config_path):
        print("⚠️  No se encontró js/supabase-config.js")
        return None, None
    with open(config_path) as f:
        content = f.read()
    import re
    url = re.search(r'SUPABASE_URL\s*=\s*["\']([^"\']+)["\']', content)
    key = re.search(r'SUPABASE_KEY\s*=\s*["\']([^"\']+)["\']', content)
    return (url.group(1) if url else None), (key.group(1) if key else None)

def fetch_supabase(url, key, corte):
    """Descarga filas de student_progress para el corte indicado."""
    sem_start, sem_end = CORTE_RANGES[corte]
    endpoint = (
        f"{url}/rest/v1/student_progress"
        f"?semana=gte.{sem_start}&semana=lte.{sem_end}"
        f"&select=student_id,student_name,grupo,semana,xp,quiz_score,activity_done"
        f"&limit=10000"
    )
    req = urllib.request.Request(endpoint, headers={
        "apikey": key, "Authorization": f"Bearer {key}"
    })
    try:
        with urllib.request.urlopen(req, timeout=20) as resp:
            return json.loads(resp.read())
    except urllib.error.URLError as e:
        print(f"❌ Error Supabase: {e}")
        sys.exit(1)

def calcular_formativa(semanas_count, actividades, quiz_scores, sem_esperadas):
    quiz_avg = sum(quiz_scores) / len(quiz_scores) if quiz_scores else 0
    vis_pct  = min(1.0, semanas_count / sem_esperadas)
    act_pct  = min(1.0, actividades / sem_esperadas)
    nota = (quiz_avg / 10 * 5.0) * W_QUIZ \
         + (act_pct * 5.0) * W_ACTIVIDADES \
         + (vis_pct * 5.0) * W_VISITAS
    return round(min(5.0, nota), 2)

def agregar_por_estudiante(rows, corte):
    """Agrupa filas Supabase por estudiante."""
    from collections import defaultdict
    bystu = defaultdict(lambda: {
        "student_id": "", "nombre": "Anónimo", "grupo": "—",
        "semanas": set(), "xp": 0, "quiz_scores": [], "actividades": 0
    })
    for r in rows:
        sid = r["student_id"]
        bystu[sid]["student_id"]  = sid
        bystu[sid]["nombre"]      = r.get("student_name") or "Anónimo"
        bystu[sid]["grupo"]       = r.get("grupo") or "—"
        bystu[sid]["semanas"].add(r["semana"])
        bystu[sid]["xp"]         += (r.get("xp") or 0)
        if r.get("quiz_score", 0) > 0:
            bystu[sid]["quiz_scores"].append(r["quiz_score"])
        if r.get("activity_done"):
            bystu[sid]["actividades"] += 1
    return dict(bystu)

def leer_parciales(path):
    """Lee CSV con columnas student_id,nota_parcial"""
    parciales = {}
    if not path or not os.path.exists(path):
        return parciales
    import csv
    with open(path, newline='', encoding='utf-8') as f:
        for row in csv.DictReader(f):
            try:
                parciales[row['student_id'].strip()] = float(row['nota_parcial'])
            except (KeyError, ValueError):
                pass
    print(f"  📄 {len(parciales)} parciales leídos desde {path}")
    return parciales

def generar_xlsx(estudiantes, parciales, corte, peso_form, output_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Corte {corte}"

    peso_parc = 1 - peso_form
    sem_start, sem_end = CORTE_RANGES[corte]
    sem_esperadas = sem_end - sem_start + 1

    # ── Estilos ────────────────────────────────────────────────────────────────
    navy_fill   = PatternFill("solid", fgColor=IUB_NAVY)
    yellow_fill = PatternFill("solid", fgColor=IUB_YELLOW)
    green_fill  = PatternFill("solid", fgColor="C8E6C9")
    red_fill    = PatternFill("solid", fgColor="FFCDD2")
    gray_fill   = PatternFill("solid", fgColor="F5F5F5")

    white_bold  = Font(bold=True, color="FFFFFF", name="Arial")
    navy_bold   = Font(bold=True, color=IUB_NAVY, name="Arial")
    yellow_bold = Font(bold=True, color="6D4C00", name="Arial")
    normal      = Font(name="Arial", size=10)
    center_al   = Alignment(horizontal="center", vertical="center")
    left_al     = Alignment(horizontal="left",   vertical="center")

    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def hcell(ws, row, col, value, fill, font, align=center_al):
        c = ws.cell(row=row, column=col, value=value)
        c.fill = fill; c.font = font; c.alignment = align; c.border = border
        return c

    # ── Encabezados ────────────────────────────────────────────────────────────
    HEADERS = [
        "#", "ID Estudiante", "Nombre", "Grupo",
        f"Semanas\nVisitadas\n(/{sem_esperadas})",
        "Actividades\nCompletas",
        "Quiz\nPromedio\n(0-10)",
        "Nota\nFormativa\n(0-5)",
        "⭐ NOTA\nPARCIAL\n(0-5)",
        f"Nota\nCorte {corte}\n(0-5)",
        "Estado"
    ]
    WIDTHS = [5, 30, 28, 10, 12, 13, 12, 14, 16, 14, 14]

    for col, (h, w) in enumerate(zip(HEADERS, WIDTHS), 1):
        fill = yellow_fill if col == 9 else navy_fill
        font = yellow_bold if col == 9 else white_bold
        c = hcell(ws, 1, col, h, fill, font)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.row_dimensions[1].height = 45

    # ── Datos ──────────────────────────────────────────────────────────────────
    for i, (sid, s) in enumerate(sorted(estudiantes.items(), key=lambda x: x[1]['nombre'])):
        nota_formativa = calcular_formativa(
            len(s["semanas"]), s["actividades"], s["quiz_scores"], sem_esperadas
        )
        parcial = parciales.get(sid)
        nota_corte = round(nota_formativa * peso_form + parcial * peso_parc, 2) if parcial is not None else None

        row = i + 2
        is_even = i % 2 == 1

        def dcell(col, val, align=center_al):
            c = ws.cell(row=row, column=col, value=val)
            c.font = normal
            c.alignment = align
            c.border = border
            if is_even and col != 9:
                c.fill = gray_fill
            return c

        dcell(1, i + 1)
        dcell(2, sid, left_al)
        dcell(3, s["nombre"], left_al)
        dcell(4, s["grupo"])
        dcell(5, len(s["semanas"]))
        dcell(6, s["actividades"])
        quiz_avg = round(sum(s["quiz_scores"]) / len(s["quiz_scores"]), 2) if s["quiz_scores"] else 0
        dcell(7, quiz_avg if quiz_avg > 0 else "Sin quiz")
        
        # Nota formativa — colorear
        c_form = dcell(8, nota_formativa)
        c_form.font = Font(bold=True, name="Arial", size=11,
                           color="1B5E20" if nota_formativa >= NOTA_MIN_APROBACION else "B71C1C")

        # Nota parcial — celda amarilla (para llenar)
        c_parc = ws.cell(row=row, column=9, value=parcial)
        c_parc.fill = PatternFill("solid", fgColor="FFF9C4")
        c_parc.font = Font(bold=True, name="Arial", size=11, color="6D4C00")
        c_parc.alignment = center_al
        c_parc.border = border

        # Nota corte — con fórmula Excel si no hay parcial
        if parcial is not None:
            c_nc = dcell(10, nota_corte)
            c_nc.font = Font(bold=True, name="Arial", size=12,
                             color="1B5E20" if nota_corte >= NOTA_MIN_APROBACION else "B71C1C")
            dcell(11, "APROBADO" if nota_corte >= NOTA_MIN_APROBACION else "REPROBADO")
        else:
            # Fórmula dinámica — se actualiza cuando el profesor llena el parcial en Excel
            h8 = get_column_letter(8); h9 = get_column_letter(9)
            c_nc = ws.cell(row=row, column=10,
                           value=f"=IF({h9}{row}=\"\",\"\",ROUND({h8}{row}*{peso_form}+{h9}{row}*{peso_parc},2))")
            c_nc.font = Font(bold=True, name="Arial", size=11, color="546E7A")
            c_nc.alignment = center_al; c_nc.border = border

            c_est = ws.cell(row=row, column=11,
                            value=f"=IF(J{row}=\"\",\"PENDIENTE\",IF(J{row}>={NOTA_MIN_APROBACION},\"APROBADO\",\"REPROBADO\"))")
            c_est.font = Font(name="Arial", size=10)
            c_est.alignment = center_al; c_est.border = border

    # ── Congelar fila de encabezados ───────────────────────────────────────────
    ws.freeze_panes = "A2"

    # ── Hoja de fórmula ────────────────────────────────────────────────────────
    wf = wb.create_sheet("Fórmula")
    info = [
        ("FÓRMULA DE CALIFICACIÓN TGA05", "", ""),
        ("", "", ""),
        ("Componente", "Peso", "Descripción"),
        ("Quiz Promedio", f"{int(W_QUIZ*100)}%", "Promedio quizzes semanas del corte (0-10 → 0-5)"),
        ("Actividades completadas", f"{int(W_ACTIVIDADES*100)}%", "Actividades / semanas esperadas"),
        ("Semanas visitadas", f"{int(W_VISITAS*100)}%", "Semanas visitadas / semanas esperadas"),
        ("", "", ""),
        ("Nota Corte", f"Formativa × {int(peso_form*100)}% + Parcial × {int(peso_parc*100)}%", ""),
        ("Nota mínima aprobación", f"{NOTA_MIN_APROBACION}", ""),
        ("Corte cubierto", f"Semanas {sem_start} a {sem_end}", ""),
        ("Fecha generación", str(date.today()), ""),
    ]
    for r, row_data in enumerate(info, 1):
        for c, val in enumerate(row_data, 1):
            cell = wf.cell(row=r, column=c, value=val)
            if r == 1:
                cell.font = Font(bold=True, size=13, color=IUB_NAVY, name="Arial")
            elif r == 3:
                cell.font = Font(bold=True, name="Arial")
                cell.fill = navy_fill
                cell.font = white_bold
    wf.column_dimensions["A"].width = 30
    wf.column_dimensions["B"].width = 35
    wf.column_dimensions["C"].width = 55

    wb.save(output_path)
    print(f"  ✅ Guardado: {output_path}")

def main():
    parser = argparse.ArgumentParser(description="Exporta notas TGA05 desde Supabase a XLSX")
    parser.add_argument("--corte", type=int, choices=[1,2,3], default=1, help="Número de corte (1, 2 o 3)")
    parser.add_argument("--peso-formativa", type=float, default=0.70, help="Peso de la nota formativa (default: 0.70)")
    parser.add_argument("--parciales", type=str, help="CSV con columnas student_id,nota_parcial")
    parser.add_argument("--output", type=str, help="Nombre del archivo de salida")
    args = parser.parse_args()

    print(f"\n📊 TGA05 — Exportar notas Corte {args.corte}")
    print(f"   Peso formativa: {int(args.peso_formativa*100)}% | Peso parcial: {int((1-args.peso_formativa)*100)}%\n")

    url, key = get_supabase_config()
    if not url or not url.startswith("http"):
        print("❌ Configura SUPABASE_URL en js/supabase-config.js")
        sys.exit(1)

    print(f"  🔗 Conectando a Supabase…")
    rows = fetch_supabase(url, key, args.corte)
    print(f"  📥 {len(rows)} filas descargadas")

    estudiantes = agregar_por_estudiante(rows, args.corte)
    print(f"  👥 {len(estudiantes)} estudiantes únicos")

    parciales = leer_parciales(args.parciales)

    output = args.output or f"TGA05_Corte{args.corte}_Notas_{date.today()}.xlsx"
    generar_xlsx(estudiantes, parciales, args.corte, args.peso_formativa, output)

if __name__ == "__main__":
    main()
